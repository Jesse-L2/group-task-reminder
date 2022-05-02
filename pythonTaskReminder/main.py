# Send emails based on task completion status in AssignmentsDue.xlsx
import os
import openpyxl
import smtplib
import ssl

# SENDER and PASSWORD are stored as environment variables (prevents sharing public credentials)
# User will need to either set these variables, or change to their personal credentials
SENDER, PASSWORD = os.environ["GMAIL_USER"], os.environ["GMAIL_PASSWORD"]

# Open spreadsheet
wb = openpyxl.load_workbook('AssignmentsDue.xlsx')
sheet = wb['Sheet1']

# complete_dict will hold the entire spreadsheet
complete_dict = {}
# Iterate through the rows of the sheet, sizing is dynamic due to max_row and max_col. Convert to a dictionary.
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column, values_only=True):
    # The index here refers to the column number, so row[0] is the employee column
    names = row[0]
    emails = row[1]

    assignments_dict = {
        # can also assign keys by sheet['C1'].value for example
        'email': emails,
        'Assignment 1': row[2],
        'Assignment 2': row[3],
        'Test 1': row[4],
        'Assignment 3': row[5],
        'Assignment 4': row[6],
        'Test 2': row[7],
        'Assignment 5': row[8],
        'Assignment 6': row[9],
        'Test 3': row[10],
        'Assignment 7': row[11],
        'Assignment 8': row[12],
        'Test 4': row[13],
    }
    complete_dict[names] = assignments_dict
# Remove the header from the print-out
del complete_dict['Student']
# incomplete_dict used to hold only students that have incomplete assignments and their incomplete assignments
incomplete_dict = {}
# Checking for None bools in the complete_dict, if they are done, we will grab those values
# Filtering out any True (complete) values
for student in complete_dict.items():
    assignments = student[1]
    for key, value in assignments.items():
        if value is None:
            # Dictionary comprehension to create a dictionary of dictionaries of incomplete assignments per person
            incomplete_dict[student[0]] = {key: value for key, value in assignments.items() if
                                           value is None or key == 'email'}


# Create an SMTP client session
# Note - if using gmail, user may have to modify some account settings within gmail itself
smtp_server = "smtp.gmail.com"
port = 587  # port will vary based on email provider, google=587, yahoo=465 for example
email = SENDER
password = PASSWORD
context = ssl.create_default_context()

try:
    with smtplib.SMTP(smtp_server, port=port) as connection:
        connection.starttls()  # SMTP connection in Transport Layer Security Mode for encrypted communications
        connection.ehlo()  # identify self to ESMTP server, not strictly required as sendmail() will call this method
        connection.login(user=SENDER, password=PASSWORD)
        # Send out reminder emails
        print(f'Incomplete dict: {incomplete_dict.items()}')
        for name, value in incomplete_dict.items():
            incomplete_assignments = []
            email = value['email']
            for assignments in value:
                incomplete_assignments.append(assignments)
            # .join() allows for removal of the [] around the list for the email for better formatting
            body = f"Dear {name},\n" \
                   "Reminder: you have the following incomplete assignments:" \
                   f" {', '.join(incomplete_assignments[1:])}" \
                   f"\nPlease turn them in at your earliest convenience."
            connection.sendmail(from_addr=SENDER, to_addrs=email, msg=body)
            print(body)
            print(f"Sending email to {email}...\n")
            sendMailStatus = connection.sendmail(SENDER, email, body)

    # TODO: Add better exception handling for more exception types
            if sendMailStatus != {}:
                print(f'There was a problem emailing {email, sendMailStatus}')
except Exception as e:
    print("There was an error!", e)
